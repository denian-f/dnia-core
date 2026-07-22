from datetime import datetime

from openpyxl import load_workbook

from app.crm.gov import config


class ExcelRepository:

    def __init__(self):

        self.wb = load_workbook(config.ARQUIVO)

    # =====================================================
    # PLANILHAS
    # =====================================================

    def clientes(self):
        return self.wb["Clientes"]

    def prospeccao(self):
        return self.wb["Prospeccao"]

    def validacao(self):
        return self.wb[config.ABA_VALIDACAO]

    # =====================================================
    # CLIENTES
    # =====================================================

    def listar_clientes(self):

        return self.clientes().iter_rows(
            min_row=2,
            values_only=True
        )

    # =====================================================
    # PROSPECÇÃO
    # =====================================================

    def listar_prospeccao(self):

        return self.prospeccao().iter_rows(
            min_row=2,
            values_only=True
        )

    def adicionar_prospeccao(self, cliente):

        agora = datetime.now().strftime("%d/%m/%Y %H:%M")

        self.prospeccao().append([

            cliente.cpf,

            cliente.nome,

            cliente.telefone,

            cliente.cidade,

            cliente.uf,

            cliente.categoria,

            cliente.posto,

            "CRM",

            agora,

            "NOVO",

            "",

            agora,

            ""

        ])

    # =====================================================
    # VALIDAÇÃO
    # =====================================================

    def listar_validacao(self):

        return self.validacao().iter_rows(
            min_row=2,
            values_only=True
        )

    def adicionar_validacao(
        self,
        cpf,
        nome
    ):

        self.validacao().append([

            cpf,

            nome,

            None

        ])

    def clientes_pendentes(self):

        clientes = []

        for indice, linha in enumerate(

            self.validacao().iter_rows(

                min_row=2,

                values_only=True

            ),

            start=2

        ):

            cpf = linha[0]
            nome = linha[1]
            gov = linha[2]

            if not cpf:
                continue

            if gov:
                continue

            clientes.append({

                "linha": indice,

                "cpf": cpf,

                "nome": nome

            })

        return clientes

    def salvar_validacao(
        self,
        linha,
        digitos
    ):

        self.validacao().cell(

            row=linha,

            column=3

        ).value = digitos

    # =====================================================
    # WORKBOOK
    # =====================================================

    def salvar(self):

        self.wb.save(config.ARQUIVO)

    def fechar(self):

        self.wb.close()